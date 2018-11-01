import csv
import openpyxl
from cataloger.models import Keyword

def keyword_import_csv(csv_file):
    decoded_file = csvfile.read().decode('utf-8').splitlines()
    reader = csv.DictReader(decoded_file)
    for row in reader:
        kw = row.split(',')[0]+row.split(',')[1]+row.split(',')[2]
        keyword = Keyword(keyword=kw)

def keyword_import_excel(excel_file):
    # open the workbook
    workbook = openpyxl.load_workbook(excel_file)
    # iterate over rows in the first sheet of the workbook
    for row in workbook[workbook.sheetnames[0]].iter_rows(row_offset=1):
        # one of the three colums won't be blank - figure out which one it is,
        # and import that into the DB as a keyword
        kw = None
        if row[0].value is not None:
            kw = str(row[0].value)
        elif row[1].value is not None:
            kw = str(row[1].value)
        elif row[2].value is not None:
            kw = str(row[2].value)
        if kw is not None:
            keyword = Keyword(keyword=kw)
            keyword.save()
